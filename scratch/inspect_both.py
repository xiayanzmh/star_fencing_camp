import openpyxl

print("=" * 60)
print("=== Aug_shenzhen.xlsx (input_data sheet) ===")
print("=" * 60)
wb1 = openpyxl.load_workbook("input_data/Aug_shenzhen.xlsx", data_only=True)
print(f"Sheets: {wb1.sheetnames}")
ws1 = wb1["input_data"]
headers1 = [cell.value for cell in ws1[1]]
print(f"Headers: {headers1}")
print(f"Total rows (including header): {ws1.max_row}")
print(f"Total cols: {ws1.max_column}")
print("\nAll rows:")
for i, row in enumerate(ws1.iter_rows(min_row=2, values_only=True)):
    if row[0] is None:
        continue
    print(f"  Row {i+2}: {row}")

print("\n" + "=" * 60)
print("=== jielong.xlsx (Sheet1) ===")
print("=" * 60)
wb2 = openpyxl.load_workbook("input_data/jielong.xlsx", data_only=True)
print(f"Sheets: {wb2.sheetnames}")
ws2 = wb2.sheetnames[0]
ws2 = wb2[ws2]
print(f"Sheet name used: {wb2.sheetnames[0]}")

# Print ALL headers
headers2 = [cell.value for cell in ws2[1]]
print(f"Headers: {headers2}")
print(f"Total rows (including header): {ws2.max_row}")
print(f"Total cols: {ws2.max_column}")

print("\nAll rows:")
for i, row in enumerate(ws2.iter_rows(min_row=2, values_only=True)):
    if all(v is None for v in row):
        continue
    print(f"  Row {i+2}: {list(row)}")
