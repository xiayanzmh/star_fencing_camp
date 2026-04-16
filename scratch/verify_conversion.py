import openpyxl

FILE = "/Users/Tommy/Desktop/python/star_fencing_camp/input_data/Aug_shenzhen_from_jielong.xlsx"

wb = openpyxl.load_workbook(FILE, data_only=True)
ws = wb["input_data"]

print(f"Total Rows: {ws.max_row}")
for i, row in enumerate(ws.iter_rows(max_row=30, values_only=True)):
    print(f"Row {i+1}: {row}")
