import openpyxl

EXCEL_FILE = "/Users/Tommy/Desktop/python/star_fencing_camp/input_data/Aug_shenzhen.xlsx"
SHEET_NAME = "input_data"

try:
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        headers = [str(cell.value).strip() if cell.value else None for cell in ws[1]]
        print(f"Headers: {headers}")
        
        # Print first few rows to see data format
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=5, values_only=True)):
            print(f"Row {i+2}: {row}")
    else:
        print(f"Sheet '{SHEET_NAME}' not found. Available sheets: {wb.sheetnames}")
except Exception as e:
    print(f"Error: {e}")
