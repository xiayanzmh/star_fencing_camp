"""
Fencing Camp Schedule Generator — Aug Shenzhen
Reads input_data from Aug_shenzhen.xlsx and generates a 6-day camp schedule.
Outputs: coach-centric CSV and kid-centric CSV.
"""
import csv
import openpyxl
from collections import defaultdict
import random
import argparse
import sys

# ─── CONFIG ──────────────────────────────────────────────────────────────
EXCEL_FILE = "Aug_shenzhen.xlsx"
SHEET_NAME = "input_data"
NUM_DAYS = 6
DAY_LABELS = [f"Day{i}" for i in range(1, NUM_DAYS + 1)]

# Sessions: (label, start_time_str, num_20min_slots)
SESSIONS = [
    ("8:00-10:00AM", "08:00", 6),
    ("1:00-2:00PM",  "13:00", 3),
    ("4:00-6:00PM",  "16:00", 6),
]

# Build ordered time-slot labels within a day
def build_time_slots():
    """Return list of time-slot labels for one day, in order."""
    slots = []
    for label, start, count in SESSIONS:
        h, m = map(int, start.split(":"))
        for i in range(count):
            total_min = h * 60 + m + i * 20
            hh = total_min // 60
            mm = total_min % 60
            slots.append(f"{hh:02d}:{mm:02d}")
    return slots

TIME_SLOTS = build_time_slots()  # 15 slots per day
SLOTS_PER_DAY = len(TIME_SLOTS)

# Morning slots (8-10 AM) indices
MORNING_SLOTS = [i for i, t in enumerate(TIME_SLOTS) if t < "12:00"]
# Afternoon slots (1-2 PM and 4-6 PM) indices
AFTERNOON_SLOTS = [i for i, t in enumerate(TIME_SLOTS) if t >= "12:00"]

# Map session start times to their slot indices
SESSION_SLOT_MAP = {}
for _label, _start, _count in SESSIONS:
    _h, _m = map(int, _start.split(":"))
    _indices = []
    for _i in range(_count):
        _total = _h * 60 + _m + _i * 20
        _slot_label = f"{_total // 60:02d}:{_total % 60:02d}"
        _idx = TIME_SLOTS.index(_slot_label)
        _indices.append(_idx)
    SESSION_SLOT_MAP[_start] = _indices

def time_pref_to_slot_indices(time_pref):
    """Convert a time_preference value (e.g. datetime.time(13,0) or '13:00') to slot indices."""
    import datetime
    if isinstance(time_pref, datetime.time):
        key = f"{time_pref.hour:02d}:{time_pref.minute:02d}"
    else:
        key = str(time_pref).strip()
        # Normalize to HH:MM
        parts = key.split(":")
        if len(parts) == 2:
            key = f"{int(parts[0]):02d}:{int(parts[1]):02d}"
    # Find the session that contains this time
    for session_start, indices in SESSION_SLOT_MAP.items():
        if key == session_start:
            return indices
    # Fallback: find closest session
    pref_min = int(key.split(":")[0]) * 60 + int(key.split(":")[1])
    best_start = None
    best_dist = float('inf')
    for session_start in SESSION_SLOT_MAP:
        sh, sm = map(int, session_start.split(":"))
        dist = abs(sh * 60 + sm - pref_min)
        if dist < best_dist:
            best_dist = dist
            best_start = session_start
    return SESSION_SLOT_MAP[best_start] if best_start else None

# ─── COACH DETECTION ───────────────────────────────────────────────────
MAIN_COACHES = []
ASST_COACHES = []
CORE_MAIN = {"吴主教练", "张杰主教练", "赵凯主教练", "Tamer主教练", "Shaimaa主教练"}
CORE_ASST = {"叶助理教练", "王助理教练"}

# ─── DATA STRUCTURES ────────────────────────────────────────────────────
schedule = {}
kid_assignments = defaultdict(list)
kid_busy = defaultdict(set)

def reset_state():
    """Clear all global scheduling data structures."""
    global schedule, kid_assignments, kid_busy
    schedule = {c: {} for c in MAIN_COACHES + ASST_COACHES}
    kid_assignments = defaultdict(list)
    kid_busy = defaultdict(set)

# ─── READ INPUT ─────────────────────────────────────────────────────────
def read_input():
    """Parse the Excel input_data sheet. Returns list of request dicts."""
    global MAIN_COACHES, ASST_COACHES
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]
    requests = []
    
    detected_main = set()
    detected_asst = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Support both 5-column and 6+ column layouts
        name = row[0]
        group = row[1] if len(row) > 1 else None
        class_num = row[2] if len(row) > 2 else None
        coach_request = row[3] if len(row) > 3 else None
        coach_type = row[4] if len(row) > 4 else None
        time_preference = row[5] if len(row) > 5 else None
        if name is None: continue
            
        if class_num is None or class_num == '-': class_num = 0
        try: class_num = int(class_num or 0)
        except: class_num = 0
            
        c_req = str(coach_request).strip() if coach_request else None
        c_type = str(coach_type).strip() if coach_type else None
        
        if c_req:
            if c_type == "主教练": detected_main.add(c_req)
            elif c_type == "助理教练": detected_asst.add(c_req)
        
        # Parse time_preference
        t_pref_indices = None
        if time_preference is not None:
            t_pref_indices = time_pref_to_slot_indices(time_preference)
            if t_pref_indices:
                print(f"  {str(name).strip()} ({c_type}): time_preference={time_preference} → slots {t_pref_indices}")
        
        requests.append({
            'name': str(name).strip(),
            'group': group,
            'class_num': class_num,
            'coach_request': c_req,
            'coach_type': c_type,
            'time_pref_slots': t_pref_indices,
        })
    
    MAIN_COACHES = sorted(list(detected_main | CORE_MAIN))
    ASST_COACHES = sorted(list(detected_asst | CORE_ASST))
    print(f"Detected Coaches: Main={MAIN_COACHES}, Asst={ASST_COACHES}")
    return requests

def read_schedule_from_excel(filename="SummerCamp_Schedule.xlsx"):
    """Parse SummerCamp_Schedule.xlsx (Day1-Day6 tabs) back into global schedule dict."""
    global schedule, kid_assignments, kid_busy
    import openpyxl
    wb = openpyxl.load_workbook(filename, data_only=True)
    reset_state()
    
    for day_idx in range(NUM_DAYS):
        day_label = DAY_LABELS[day_idx]
        if day_label not in wb.sheetnames:
            print(f"  Warning: Sheet {day_label} not found in {filename}")
            continue
        ws = wb[day_label]
        
        # Load headers to find coach columns
        headers = [str(cell.value).strip() if cell.value else None for cell in ws[1]]
        coach_map = {} # col_idx -> coach_name
        for i, h in enumerate(headers):
            if h and h in (MAIN_COACHES + ASST_COACHES):
                coach_map[i+1] = h
        
        # Iterate rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            slot_label = str(row[0]).strip() if row[0] else None
            # Check if this is a time slot (HH:MM)
            if slot_label and ":" in slot_label and len(slot_label) <= 5:
                try:
                    slot_idx = TIME_SLOTS.index(slot_label)
                except ValueError:
                    continue
                
                for col_idx, coach in coach_map.items():
                    kid = row[col_idx-1]
                    if kid and str(kid).strip():
                        kid_name = str(kid).strip()
                        assign_slot(kid_name, coach, day_idx, slot_idx)
    print(f"  Successfully read {sum(len(s) for s in schedule.values())} assignments from {filename}")

# ─── SCHEDULING HELPERS ─────────────────────────────────────────────────
def is_slot_free(coach, day_idx, slot_idx):
    return (day_idx, slot_idx) not in schedule[coach]

def assign_slot(kid, coach, day_idx, slot_idx):
    schedule[coach][(day_idx, slot_idx)] = kid
    kid_assignments[kid].append((day_idx, slot_idx, coach))
    kid_busy[kid].add((day_idx, slot_idx))

def get_coach_load(coach):
    return len(schedule[coach])

def get_coach_day_load(coach, day_idx):
    return sum(1 for (d, s) in schedule[coach] if d == day_idx)

def get_prioritized_candidates(shuffle_slots=False, priority_indices=None):
    morning = [i for i in MORNING_SLOTS]
    early_as = [i for i, t in enumerate(TIME_SLOTS) if "13:00" <= t < "14:00"]
    late_as = [i for i, t in enumerate(TIME_SLOTS) if t >= "16:00"]
    
    if shuffle_slots:
        random.shuffle(morning)
        random.shuffle(early_as)
        random.shuffle(late_as)
    
    all_slots = morning + early_as + late_as
    
    if priority_indices:
        # Create a copy to avoid modifying globals if passed directly
        p_list = list(priority_indices)
        # Filter all_slots to get those not in priority
        others = [i for i in all_slots if i not in p_list]
        return p_list + others
    return all_slots

def is_time_separated(kid, day_idx, slot_idx):
    """Ensure kid's second session on a day is in a different block (AM/PM)."""
    existing = [(d, s) for (d, s) in kid_busy[kid] if d == day_idx]
    if not existing: return True
    
    existing_slot = existing[0][1]
    existing_is_morning = existing_slot in MORNING_SLOTS
    new_is_morning = slot_idx in MORNING_SLOTS
    new_is_afternoon = slot_idx in AFTERNOON_SLOTS
    
    if existing_is_morning: return new_is_afternoon
    else: return new_is_morning

def find_consistent_multi_coach_slots(kid, coach_list, num_needed, shuffle_slots=False, preferred_slot_indices=None, locked_coach=None, is_dual=False):
    candidates = get_prioritized_candidates(shuffle_slots, preferred_slot_indices)
    for slot_idx in candidates:
        available_options = []
        day_indices = list(range(NUM_DAYS))
        if shuffle_slots: random.shuffle(day_indices)
        
        for day_idx in day_indices:
            if (day_idx, slot_idx) in kid_busy[kid]: continue
            if is_dual and not is_time_separated(kid, day_idx, slot_idx): continue
                
            if locked_coach:
                if is_slot_free(locked_coach, day_idx, slot_idx):
                    available_options.append((day_idx, locked_coach))
            else:
                coaches_sorted = sorted(coach_list, key=lambda c: (get_coach_day_load(c, day_idx), get_coach_load(c)))
                for coach in coaches_sorted:
                    if is_slot_free(coach, day_idx, slot_idx):
                        available_options.append((day_idx, coach))
                        break
        
        if len(available_options) >= num_needed:
            available_options.sort(key=lambda x: get_coach_day_load(x[1], x[0]))
            chosen = available_options[:num_needed]
            chosen.sort(key=lambda x: x[0])
            return [(d, slot_idx, c) for d, c in chosen]
    return None

def find_best_slots_flexible_v2(kid, coach_list, num_needed, shuffle_slots=False, preferred_slot_indices=None, locked_coach=None, is_dual=False):
    result = find_consistent_multi_coach_slots(kid, coach_list, num_needed, shuffle_slots, preferred_slot_indices, locked_coach, is_dual)
    if result: return result
    
    candidates = get_prioritized_candidates(shuffle_slots, preferred_slot_indices)
    available = []
    for slot_idx in candidates:
        day_indices = list(range(NUM_DAYS))
        if shuffle_slots: random.shuffle(day_indices)
        for day_idx in day_indices:
            if (day_idx, slot_idx) in kid_busy[kid]: continue
            if is_dual and not is_time_separated(kid, day_idx, slot_idx): continue
            
            if locked_coach:
                if is_slot_free(locked_coach, day_idx, slot_idx):
                    available.append((day_idx, slot_idx, locked_coach))
            else:
                coaches_sorted = sorted(coach_list, key=lambda c: get_coach_load(c))
                for coach in coaches_sorted:
                    if is_slot_free(coach, day_idx, slot_idx):
                        available.append((day_idx, slot_idx, coach))
                        break
    if len(available) >= num_needed: return available[:num_needed]
    return None

# ─── MAIN SCHEDULING LOGIC ─────────────────────────────────────────────
def build_schedule(shuffle_slots=False):
    requests = read_input()
    reset_state()
    
    main_requests = [r for r in requests if r['coach_type'] == '主教练']
    asst_requests = [r for r in requests if r['coach_type'] == '助理教练']
    main_kids = set(r['name'] for r in main_requests)
    asst_kids = set(r['name'] for r in asst_requests)
    dual_kids = main_kids & asst_kids
    
    # ── PHASE 1: Coach-requested 主教练 ──────
    print(f"=== Phase 1: Coach-requested 主教练 (Shuffle={shuffle_slots}) ===")
    requested_main = [r for r in main_requests if r['coach_request']]
    requested_main.sort(key=lambda r: (r['name'] not in dual_kids, -r['class_num']))
    for req in requested_main:
        kid, coach, num = req['name'], req['coach_request'], req['class_num']
        if coach not in MAIN_COACHES: continue
        pref_slots = req.get('time_pref_slots') or MORNING_SLOTS
        slots = find_best_slots_flexible_v2(kid, [coach], num, shuffle_slots, preferred_slot_indices=pref_slots, locked_coach=coach, is_dual=kid in dual_kids)
        if slots:
            for d, s, c in slots: assign_slot(kid, c, d, s)
    
    # ── PHASE 2: Coach-requested 助理教练 ────
    print(f"\n=== Phase 2: Coach-requested 助理教练 (Shuffle={shuffle_slots}) ===")
    requested_asst = [r for r in asst_requests if r['coach_request']]
    requested_asst.sort(key=lambda r: (r['name'] not in dual_kids, -r['class_num']))
    for req in requested_asst:
        kid, coach, num = req['name'], req['coach_request'], req['class_num']
        if coach not in ASST_COACHES: continue
        pref_slots = req.get('time_pref_slots') or AFTERNOON_SLOTS
        slots = find_best_slots_flexible_v2(kid, [coach], num, shuffle_slots, preferred_slot_indices=pref_slots, locked_coach=coach, is_dual=kid in dual_kids)
        if slots:
            for d, s, c in slots: assign_slot(kid, c, d, s)
    
    # ── PHASE 3: No-preference 主教练 ──────
    print(f"\n=== Phase 3: No-preference 主教练 (Shuffle={shuffle_slots}) ===")
    no_pref_main = [r for r in main_requests if not r['coach_request']]
    no_pref_main.sort(key=lambda r: (r['name'] not in dual_kids, -r['class_num']))
    for req in no_pref_main:
        kid, num = req['name'], req['class_num']
        pref_slots = req.get('time_pref_slots') or MORNING_SLOTS
        slots = find_best_slots_flexible_v2(kid, MAIN_COACHES, num, shuffle_slots, preferred_slot_indices=pref_slots, is_dual=kid in dual_kids)
        if slots:
            for d, s, c in slots: assign_slot(kid, c, d, s)
    
    # ── PHASE 4: No-preference 助理教练 ────
    print(f"\n=== Phase 4: No-preference 助理教练 (Shuffle={shuffle_slots}) ===")
    no_pref_asst = [r for r in asst_requests if not r['coach_request']]
    no_pref_asst.sort(key=lambda r: (r['name'] not in dual_kids, -r['class_num']))
    for req in no_pref_asst:
        kid, num = req['name'], req['class_num']
        pref_slots = req.get('time_pref_slots') or AFTERNOON_SLOTS
        slots = find_best_slots_flexible_v2(kid, ASST_COACHES, num, shuffle_slots, preferred_slot_indices=pref_slots, is_dual=kid in dual_kids)
        if slots:
            for d, s, c in slots: assign_slot(kid, c, d, s)
    
    return requests

# ─── OUTPUT ─────────────────────────────────────────────────────────────
def write_coach_csv(filename="schedule_by_coach.csv"):
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        for coach in MAIN_COACHES + ASST_COACHES:
            coach_schedule = schedule[coach]
            if not coach_schedule: continue
            writer.writerow([])
            writer.writerow([f"=== {coach} === (Total: {len(coach_schedule)} lessons)"])
            writer.writerow(["Time Slot"] + DAY_LABELS)
            for slot_idx, time_label in enumerate(TIME_SLOTS):
                row = [time_label]
                for day_idx in range(NUM_DAYS):
                    row.append(coach_schedule.get((day_idx, slot_idx), ""))
                writer.writerow(row)

def write_kid_csv(filename="schedule_by_kid.csv"):
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(["Kid Name", "Day", "Time", "Coach", "Coach Type"])
        for kid in sorted(kid_assignments.keys()):
            lessons = sorted(kid_assignments[kid], key=lambda x: (x[0], x[1]))
            for day_idx, slot_idx, coach in lessons:
                ctype = "主教练" if coach in MAIN_COACHES else "助理教练"
                writer.writerow([kid, DAY_LABELS[day_idx], TIME_SLOTS[slot_idx], coach, ctype])

def write_summary_csv(requests, filename="schedule_summary.csv"):
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(["Kid Name", "Coach Type", "Requested Coach", "Expected", "Assigned", "Coaches", "Status"])
        for req in requests:
            kid, coach_type, expected = req['name'], req['coach_type'], req['class_num']
            count = sum(1 for d, s, c in kid_assignments.get(kid, []) if ("主教练" if c in MAIN_COACHES else "助理教练") == coach_type)
            actual_coaches = set(c for d, s, c in kid_assignments.get(kid, []) if ("主教练" if c in MAIN_COACHES else "助理教练") == coach_type)
            status = "✓" if count >= expected else f"SHORT by {expected - count}"
            writer.writerow([kid, coach_type, req['coach_request'] or "Any", expected, count, ", ".join(actual_coaches), status])

def write_comparison_csv(requests, a1, a2, filename="schedule_comparison.csv"):
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(["Kid Name", "Coach Type", "V1 Assigned", "V2 Assigned", "Diff"])
        for req in requests:
            kid, ctype = req['name'], req['coach_type']
            def fmt(lessons, ct):
                filtered = sorted([l for l in lessons if ("主教练" if l[2] in MAIN_COACHES else "助理教练") == ct], key=lambda x: (x[0],x[1]))
                return "; ".join([f"{DAY_LABELS[d]} {TIME_SLOTS[s]} ({c})" for d, s, c in filtered])
            v1_s = fmt(a1.get(kid, []), ctype)
            v2_s = fmt(a2.get(kid, []), ctype)
            writer.writerow([kid, ctype, v1_s, v2_s, "Same" if v1_s == v2_s else "Shifted"])

def write_json(filename="schedule.json"):
    import json
    data = {"coaches": MAIN_COACHES + ASST_COACHES, "days": DAY_LABELS, "slots": TIME_SLOTS, "assignments": []}
    for coach in data["coaches"]:
        ctype = "主教练" if coach in MAIN_COACHES else "助理教练"
        for (day_idx, slot_idx), kid in schedule[coach].items():
            data["assignments"].append({"kid": kid, "coach": coach, "coach_type": ctype, "day": day_idx, "slot": slot_idx})
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def write_excel(filename="SummerCamp_Schedule.xlsx"):
    """Write a multi-tab Excel file with one sheet per day (Day1–Day6).
    Each sheet: rows = time slots, columns = coaches."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    all_coaches = MAIN_COACHES + ASST_COACHES
    
    # Style definitions
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    session_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    session_font = Font(bold=True, size=10, color="2F5496")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    # Session boundaries for visual grouping
    session_ranges = []
    for label, start, count in SESSIONS:
        h, m = map(int, start.split(":"))
        first_slot = f"{h:02d}:{m:02d}"
        first_idx = TIME_SLOTS.index(first_slot)
        session_ranges.append((label, first_idx, first_idx + count - 1))
    
    for day_idx in range(NUM_DAYS):
        ws = wb.create_sheet(title=DAY_LABELS[day_idx])
        
        # Header row: Time Slot | Coach1 | Coach2 | ...
        ws.cell(row=1, column=1, value="Time Slot").font = header_font_white
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = cell_alignment
        ws.cell(row=1, column=1).border = thin_border
        ws.column_dimensions['A'].width = 14
        
        for col_idx, coach in enumerate(all_coaches, start=2):
            cell = ws.cell(row=1, column=col_idx, value=coach)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = cell_alignment
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = 22
        
        # Data rows: one per time slot
        current_row = 2
        for sess_label, sess_start, sess_end in session_ranges:
            # Session header row
            cell = ws.cell(row=current_row, column=1, value=sess_label)
            cell.font = session_font
            cell.fill = session_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
            for col_idx in range(2, len(all_coaches) + 2):
                c = ws.cell(row=current_row, column=col_idx)
                c.fill = session_fill
                c.border = thin_border
            current_row += 1
            
            for slot_idx in range(sess_start, sess_end + 1):
                time_label = TIME_SLOTS[slot_idx]
                ws.cell(row=current_row, column=1, value=time_label).alignment = cell_alignment
                ws.cell(row=current_row, column=1).border = thin_border
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                
                for col_idx, coach in enumerate(all_coaches, start=2):
                    kid = schedule[coach].get((day_idx, slot_idx), "")
                    cell = ws.cell(row=current_row, column=col_idx, value=kid)
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                current_row += 1
        
        # Freeze top row
        ws.freeze_panes = "B2"
    
    wb.save(filename)
    print(f"  Excel saved: {filename}")

def write_coach_excel(filename="schedule_by_coach.xlsx"):
    """Write a multi-tab Excel where each tab is one coach.
    Columns = Day1 to Day6, Rows = Time Slots."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    wb.remove(wb.active)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    session_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    session_font = Font(bold=True, size=10, color="2F5496")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    # Session boundaries for visual grouping
    session_ranges = []
    for label, start, count in SESSIONS:
        h, m = map(int, start.split(":"))
        first_slot = f"{h:02d}:{m:02d}"
        first_idx = TIME_SLOTS.index(first_slot)
        session_ranges.append((label, first_idx, first_idx + count - 1))
        
    all_coaches = MAIN_COACHES + ASST_COACHES
    for coach in all_coaches:
        coach_sched = schedule[coach]
        if not coach_sched and coach not in CORE_MAIN and coach not in CORE_ASST: continue
        
        ws = wb.create_sheet(title=coach[:31]) # Excel sheet name limit
        
        # Header Row
        ws.cell(row=1, column=1, value="Time Slot").font = header_font_white
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).border = thin_border
        ws.column_dimensions['A'].width = 14
        
        for d_idx, label in enumerate(DAY_LABELS):
            cell = ws.cell(row=1, column=d_idx+2, value=label)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = cell_alignment
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = 18
            
        # Data Rows with Session Headers
        current_row = 2
        for sess_label, sess_start, sess_end in session_ranges:
            # Session header row
            cell = ws.cell(row=current_row, column=1, value=sess_label)
            cell.font = session_font
            cell.fill = session_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
            for col_idx in range(2, NUM_DAYS + 2):
                c = ws.cell(row=current_row, column=col_idx)
                c.fill = session_fill
                c.border = thin_border
            current_row += 1
            
            for s_idx in range(sess_start, sess_end + 1):
                slot_label = TIME_SLOTS[s_idx]
                ws.cell(row=current_row, column=1, value=slot_label).border = thin_border
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                ws.cell(row=current_row, column=1).alignment = cell_alignment
                
                for d_idx in range(NUM_DAYS):
                    val = coach_sched.get((d_idx, s_idx), "")
                    cell = ws.cell(row=current_row, column=d_idx+2, value=val)
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                current_row += 1
                
        ws.freeze_panes = "B2"
        
    wb.save(filename)
    print(f"  Coach Excel saved: {filename}")

def validate_against_input(requests):
    """Compare assignment counts against the original requests and print a summary."""
    print("\n" + "="*50)
    print("VALIDATION REPORT: Requested vs Assigned")
    print("="*50)
    
    # Aggregate requests: (kid, coach_type) -> total_expected
    aggregated_req = defaultdict(int)
    for r in requests:
        aggregated_req[(r['name'], r['coach_type'])] = aggregated_req[(r['name'], r['coach_type'])] + r['class_num']
        
    errors = 0
    total_requested = 0
    total_assigned_count = 0
    
    # We also track which kids were assigned but NOT requested
    all_assigned_kids = set(kid_assignments.keys())
    all_requested_keys = set(aggregated_req.keys())
    
    for (kid, ctype), expected in sorted(aggregated_req.items()):
        total_requested += expected
        # Count actual from rebuilt state. Note: kid_assignments[kid] is list of (day, slot, coach)
        actual = sum(1 for d, s, c in kid_assignments[kid] if ("主教练" if c in MAIN_COACHES else "助理教练") == ctype)
        
        if actual < expected:
            print(f"  SHORT: {kid} ({ctype}) -> Req:{expected}, Got:{actual} [Diff: {expected-actual}]")
            errors += 1
        elif actual > expected:
            print(f"  OVER : {kid} ({ctype}) -> Req:{expected}, Got:{actual} [Diff: {actual-expected}]")
            errors += 1
            
    # Check for kids present in schedule but not in input
    for kid in sorted(all_assigned_kids):
        actual_total = len(kid_assignments[kid])
        total_assigned_count += actual_total
        main_actual = sum(1 for d, s, c in kid_assignments[kid] if c in MAIN_COACHES)
        asst_actual = sum(1 for d, s, c in kid_assignments[kid] if c in ASST_COACHES)
        
        if main_actual > 0 and (kid, "主教练") not in all_requested_keys:
            print(f"  UNEXPECTED: {kid} (主教练) -> Assigned:{main_actual} but no request found.")
            errors += 1
        if asst_actual > 0 and (kid, "助理教练") not in all_requested_keys:
            print(f"  UNEXPECTED: {kid} (助理教练) -> Assigned:{asst_actual} but no request found.")
            errors += 1
            
    print("-"*50)
    print(f"Summary: Total Requested={total_requested}, Total Assigned={total_assigned_count}")
    if errors == 0:
        print("✓ All counts match perfectly!")
    else:
        print(f"✗ Found {errors} discrepancies.")
    print("="*50 + "\n")
    return errors

def validate():
    # Keep the basic coach load summary
    print("\n=== Coach Load Summary ===")
    for coach in MAIN_COACHES + ASST_COACHES:
        total = len(schedule[coach])
        by_day = defaultdict(int)
        for d, s in schedule[coach]: by_day[d] += 1
        day_str = ", ".join(f"D{d+1}:{by_day[d]}" for d in range(NUM_DAYS))
        print(f"  {coach}: {total} total ({day_str})")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Fencing Camp Schedule Generator")
    parser.add_argument("--step1", action="store_true", help="Generate initial SummerCamp_Schedule.xlsx from Aug_shenzhen.xlsx")
    parser.add_argument("--step2", action="store_true", help="Generate final coach/kid reports from SummerCamp_Schedule.xlsx (respects manual edits)")
    args = parser.parse_args()

    random.seed(42)
    
    if not args.step1 and not args.step2:
        print("Please specify --step1 or --step2. (Example: python3 schedule_generator.py --step1)")
        sys.exit(0)

    if args.step1:
        print("\n>>> STEP 1: GENERATING INITIAL SCHEDULE <<<")
        requests = build_schedule(shuffle_slots=False)
        validate()
        
        write_json("schedule.json")
        write_coach_csv("schedule_by_coach.csv") # legacy CSV still useful
        write_kid_csv("schedule_by_kid.csv")
        write_summary_csv(requests, "schedule_summary.csv")
        write_excel("SummerCamp_Schedule.xlsx")
        print("\nStep 1 Complete! You can now edit SummerCamp_Schedule.xlsx if needed.")

    if args.step2:
        print("\n>>> STEP 2: GENERATING FINAL REPORTS FROM EXCEL <<<")
        # We still need to read input to know requested counts and coach types
        requests = read_input()
        
        # Read the (potentially edited) schedule back from Excel
        read_schedule_from_excel("SummerCamp_Schedule.xlsx")
        
        # Sync the JSON and CSVs from the current state (parsed from Excel)
        write_json("schedule.json")
        write_coach_excel("schedule_by_coach.xlsx")
        write_kid_csv("schedule_by_kid.csv")
        
        # Final validation against input
        validate_against_input(requests)
        validate() # print load summary
        print("\nStep 2 Complete! Generated schedule_by_coach.xlsx, schedule_by_kid.csv and updated schedule.json")
