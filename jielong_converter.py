import openpyxl
import re
import os
import argparse
from openpyxl import Workbook

# CONFIG
INPUT_FILE = "input_data/jielong.xlsx"
OUTPUT_FILE = "input_data/Aug_shenzhen_from_jielong.xlsx"

# Standard coach names from schedule_generator.py
MAIN_COACHES = ["吴主教练", "张杰主教练", "赵凯主教练", "Tamer主教练", "Shaimaa主教练"]
ASST_COACHES = ["叶助理教练", "王助理教练"]

def clean_text(text):
    if not text: return ""
    # Normalize common Chinese punctuation to English
    text = str(text).replace("，", ",").replace("（", "(").replace("）", ")").replace("：", ":")
    # Remove leading/trailing numbers like "1. ", "2) " often found in jielong lists
    text = re.sub(r"^\d+[\.\)\s、]+", "", text)
    return text.strip()

def clean_name(name):
    # Remove common artifacts from name extraction
    name = re.sub(r"\(|\)|-|—|：|:|,$", "", name).strip()
    return name

CHINESE_NUM_MAP = {
    "一": 1, "二": 2, "两": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9, "十": 10, "十一": 11, "十二": 12
}

def parse_num(s):
    if s.isdigit(): return int(s)
    return CHINESE_NUM_MAP.get(s, 0)

def detect_group(text):
    if "男花" in text: return "男花"
    if "女花" in text: return "女花"
    if "花剑" in text: return "花剑"
    return "-"

def map_coach_request(text):
    if not text: return None
    if "吴" in text: return "吴主教练"
    if "张" in text: return "张杰主教练"
    if "赵" in text or "Kai" in text: return "赵凯主教练"
    if "Tamer" in text: return "Tamer主教练"
    if "Shaimaa" in text: return "Shaimaa主教练"
    if "叶" in text: return "叶助理教练"
    if "王" in text: return "王助理教练"
    return None

def parse_jielong(max_main_lessons=None):
    if not os.path.exists(INPUT_FILE):
        print(f"Error: {INPUT_FILE} not found.")
        return

    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws = wb.active # Usually Sheet1
    
    records = []
    
    # Iterate through rows, skipping header
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        # Join all text in row to handle split cells (e.g. Row 89)
        all_text = " ".join([str(c) for c in row[1:] if c is not None])
        raw_text = clean_text(all_text)
        if not raw_text: continue
        
        # --- SPLIT STRATEGY ---
        # Handle "A & B 各 X 节" or "A, B, C 各 X 节"
        is_each = "各" in raw_text
        
        # 1. Identify Students
        # Search for names/groups before the first "节" or "各" or "主教练" etc.
        # Added coach names to stop regex to avoid them bleeding into student names
        m_stop = re.search(r"(\d+|[一二三四五六七八九十]|各)节|各|主教练|助理教练|主教|助教|吴|张|赵|Kai|Tamer|Shaimaa|叶|王", raw_text)
        if m_stop:
            student_part = raw_text[:m_stop.start()]
        else:
            student_part = raw_text
            
        # Split names by common delimiters
        names = [clean_name(n) for n in re.split(r"[&+、,和]", student_part) if clean_name(n)]
        if not names:
            # Fallback: take the first word
            names = [clean_name(raw_text.split()[0])]

        # 2. Extract Requests (count + coach type)
        # Search for patterns like "6节(主教练)" or "主教4节"
        is_40min = "40分钟" in raw_text

        # If slash is present with "节" on both sides, it might be a clarification (X / Y)
        segments = [raw_text]
        if "/" in raw_text:
            segments = [s.strip() for s in raw_text.split("/") if "节" in s]
            if len(segments) > 1:
                with_40 = [s for s in segments if "40分钟" in s]
                if with_40: segments = [with_40[0]]
                else: segments = [segments[0]]

        found_requests = []
        for seg in segments:
            # Pattern 1: (count)节 ... (coach)
            # Added (?:私[课教])? to handle "6节私课"
            matches = re.findall(r"(\d+|[一二三四五六七八九十])节(?:私[课教])?\s*\(?(主教练|助理教练|主教|助教|吴|张|赵|Kai|Tamer|Shaimaa|叶|王)?\)?", seg)
            for count_str, c_info in matches:
                count = parse_num(count_str)
                if count == 0: continue
                if "40分钟" in seg: count *= 2
                
                c_type = "主教练"
                if c_info and ("助" in c_info or "叶" in c_info or "王" in c_info):
                    c_type = "助理教练"
                
                c_req = map_coach_request(c_info)
                found_requests.append({'count': count, 'type': c_type, 'request': c_req})

            # Pattern 2: (coach) ... (count)节
            if not matches:
                matches = re.findall(r"(主教练|助理教练|主教|助教|吴|张|赵|Kai|Tamer|Shaimaa|叶|王)\s*(?:私[课教])?(\d+|[一二三四五六七八九十])节", seg)
                for c_info, count_str in matches:
                    count = parse_num(count_str)
                    if count == 0: continue
                    if "40分钟" in seg: count *= 2
                    c_type = "主教练"
                    if c_info and ("助" in c_info or "叶" in c_info or "王" in c_info):
                        c_type = "助理教练"
                    c_req = map_coach_request(c_info)
                    found_requests.append({'count': count, 'type': c_type, 'request': c_req})

        # Deduplicate requests
        unique_reqs = []
        seen = set()
        for r in found_requests:
            key = (r['count'], r['type'], r['request'])
            if key not in seen:
                unique_reqs.append(r)
                seen.add(key)
        
        # Fallback if no specific count found but name exists
        if not unique_reqs and names:
            unique_reqs.append({'count': 6, 'type': '主教练', 'request': None})

        # Apply Max Main Lessons restriction if set
        if max_main_lessons:
            for req in unique_reqs:
                if req['type'] == '主教练' and req['count'] > max_main_lessons:
                    print(f"  NOTICE: Capping {names[0] if names else 'unknown'} (主教练) from {req['count']} to {max_main_lessons} sessions.")
                    req['count'] = max_main_lessons
        
        # If "各" is present, we apply the requests to ALL names
        # Otherwise, if we have multiple names and multiple requests, it's tricky.
        # Usually it's "Student1 + Student2 各 X 节" OR "Student1 request1, Student2 request2"
        
        if is_each:
            for name in names:
                for req in unique_reqs:
                    records.append({
                        'Name': name,
                        'group': detect_group(raw_text),
                        'class_num': req['count'],
                        'coach_request': req['request'],
                        'coach_type': req['type']
                    })
        else:
            # Handle specific multi-student rows like Row 43 (Eason + Mira)
            # If we have 2 names and 2 requests, pair them?
            # Actually, most rows have 1 student or clear "each" logic.
            # If we have 1 name and multiple requests (e.g. Row 42: 主教3, 助教3)
            if len(names) == 1:
                for req in unique_reqs:
                    records.append({
                        'Name': names[0],
                        'group': detect_group(raw_text),
                        'class_num': req['count'],
                        'coach_request': req['request'],
                        'coach_type': req['type']
                    })
            else:
                # Basic heuristic: if no "each", assume requests apply to first name or attempt to find name in text
                for name in names:
                    # Look for specific mentions of this name near a count (hard)
                    # For now, just apply ALL requests if they seem distinctive
                    # E.g. "John 6, Lucas 4" -> we should ideally split.
                    # Let's try to find if the name is specifically before a count.
                    for req in unique_reqs:
                        # Check if this request is relevant to this name
                        # (Very basic: if name is before count)
                        records.append({
                            'Name': name,
                            'group': detect_group(raw_text),
                            'class_num': req['count'],
                            'coach_request': req['request'],
                            'coach_type': req['type']
                        })

    # --- WRITE OUTPUT ---
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "input_data"
    
    headers = ["Name", "group", "class_num", "coach_request", "coach_type", "time_preference"]
    out_ws.append(headers)
    
    # Deduplicate final records by (Name, type, coach)
    # Actually, some kids might have multiple requests that should sum up, 
    # but the generator prefers them separate if coaches differ.
    
    for r in records:
        out_ws.append([
            r['Name'],
            r['group'],
            r['class_num'],
            r['coach_request'],
            r['coach_type'],
            None # time_preference
        ])
        
    out_wb.save(OUTPUT_FILE)
    print(f"Successfully converted {len(records)} records to {OUTPUT_FILE}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert jielong.xlsx to structured format")
    parser.add_argument("--max-main", type=int, help="Max lessons allowed for a Main Coach request")
    args = parser.parse_args()

    m_main = args.max_main
    if m_main is None:
        try:
            val = input("Enter max class_num for coach_type = '主教练' (default 4, press Enter to skip): ").strip()
            if val:
                m_main = int(val)
            else:
                m_main = 4 # Default if they just press enter
        except (ValueError, EOFError):
            m_main = 4
            print("Using default 4.")

    print(f"\n>>> Converting jielong with MAX_MAIN_LESSONS={m_main} <<<")
    parse_jielong(max_main_lessons=m_main)
